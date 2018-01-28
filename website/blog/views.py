# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate
from django.contrib.auth import logout as django_logout
from django.contrib.auth import login as django_login
from django.contrib.auth.models import User
from django.http import Http404, HttpResponse
from .models import Post, UserProfile, Comment
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import datetime
import openpyxl as opxl
from openpyxl.writer.excel import save_virtual_workbook
# Create your views here.


def home(request):
    if request.user.is_authenticated():
        return redirect('blog:userpage')
    else:
        return redirect('blog:login')

#Sign-Up, Login and Logout

def signup(request):
    #Displays sign-up page
    if request.method == 'GET':
        return render(request, 'blog/signup.html')
    #Signs up the user
    elif request.method =='POST':
        fname = request.POST['fname']
        lname = request.POST['lname']
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password_re = request.POST['password_re']


        if fname.isalpha() and lname.isalpha() and username.isalnum():
            if password == password_re:
                #Creates User
                user = User(username = username, email = email, first_name = fname, last_name = lname)
                user.set_password(password)
                user.save()
                #Creates User Profile
                userprofile = UserProfile(user=user)
                userprofile.save()

                return render(request, 'blog/login.html')

            else:
                messages.error(request, "Your passwords don't match!")
                return redirect('blog:signup')

        else:
            messages.error(request, "Please fill all the fields!")
            return redirect('blog:signup')



def login(request):
    #Display the login page
    if request.method == 'GET':
        return render(request, 'blog/login.html')
    #Logging user into the website
    elif request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        print("username: {}, password: {}".format(username, password))
        user = authenticate(username = username, password = password)
        print user

        if user is not None:
            django_login(request, user)
            messages.success(request, "Successful Login")
            return redirect('blog:userpage')
        else:
            messages.error(request, "Please enter the correct credentials and try again!")
            return redirect('blog:login')

def logout(request):
    django_logout(request)
    return redirect('blog:home')

#Following views are only for logged in users

@login_required
def userpage(request):
    my_user = request.user
    return render(request, 'blog/userpage.html', {'user':my_user,})

@login_required
def post_detail(request, post_pk):
    post = get_object_or_404(Post, pk = post_pk)
    comments = post.comment_set.all()
    return render(request, 'blog/post_detail.html', {'post':post, 'comments':comments,})

@login_required
def userlist(request):
    my_user = request.user
    userlist = User.objects.exclude(pk = my_user.pk)
    return render(request, 'blog/userlist.html', {'userlist':userlist})

@login_required
def user_follow(request, username):
    my_user =  request.user
    user = User.objects.get(username = username)
    my_userprofile = my_user.user_profile
    my_userprofile.follow.add(user)
    return redirect('blog:following')

@login_required
def following(request):
    my_user = request.user
    userprofile = my_user.user_profile
    follow_list = userprofile.follow.all()
    return render(request, 'blog/following.html', {'follow_list':follow_list,})

@login_required
def user_home(request):
    my_user = request.user
    follow_list = my_user.user_profile.follow.all()
    post_list = list()
    for user in follow_list:
        for post in user.post_set.all():
            post_list.append(post)
    for i in range(0, len(post_list)-1):
        for j in range(i+1, len(post_list)):
            if post_list[i].pub_date < post_list[j].pub_date:
                post_list[i], post_list[j] = post_list[j], post_list[i]
    return render(request, 'blog/user_home.html', {'post_list':post_list})

@login_required
def user_posts(request):
    my_user = request.user
    posts = my_user.post_set.all()
    return render(request, 'blog/user_posts.html', {'posts':posts})

@login_required
def write_post(request):
    if request.method == 'GET':
        return render(request, 'blog/write_post.html')
    elif request.method == 'POST':
        my_user = request.user
        title = request.POST['title']
        content = request.POST['content']
        pub_date = datetime.datetime.now()
        post = Post(user = my_user, title = title, content = content)
        post.save()
        return redirect('blog:post_detail', post_pk = post.pk)

@login_required
def comment(request, post_pk):
    my_user = request.user
    post = Post.objects.get(pk = post_pk)
    comment_text = request.POST['comment']
    comment = Comment(user = my_user, post = post, comment = comment_text)
    comment.save()
    return post_detail(request, post_pk)

@login_required
def delete_post(request, post_pk):
    post = Post.objects.get(pk = post_pk)
    post.delete()
    return userpage(request, request.user.username)


def export_xlsx(request):
    user_list = User.objects.exclude(username = 'admin')
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = 'User Information'

    ws['A1'] = "Primary Key"
    ws['B1'] = "First Name"
    ws['C1'] = "Last Name"
    ws['D1'] = "Username"
    ws['E1'] = "Email"
    ws['F1'] = "DOB"
    ws['G1'] = "Mobile No"

    for user in user_list:
        ws['A{}'.format(user.pk+1)] = user.pk
        ws['B{}'.format(user.pk+1)] = user.first_name
        ws['C{}'.format(user.pk+1)] = user.last_name
        ws['D{}'.format(user.pk+1)] = user.username
        ws['E{}'.format(user.pk+1)] = user.email
        ws['F{}'.format(user.pk+1)] = user.user_profile.dob
        ws['G{}'.format(user.pk+1)] = user.user_profile.mobile_no

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename = "user_info.xlsx"'
    return response
